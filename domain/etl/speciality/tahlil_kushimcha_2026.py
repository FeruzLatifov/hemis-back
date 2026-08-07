#!/usr/bin/env python3
"""
Tahlil hisoboti: Бакалавр/Магистр-кушимча-2026.xlsx dagi 66 qatorning har biri
bazada (S014/S015 = etl_speciality*.csv) BOR-mi yoki YO'Q-mi — dalillar bilan.
Chiqish: TAHLIL_kushimcha_2026.xlsx (2 varaq: Xulosa + Tahlil).
"""
import openpyxl, csv, re, unicodedata
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APOS = "'’ʻʼ‘`"
M = {'ё':'yo','ж':'j','ц':'ts','щ':'sh','ю':'yu','я':'ya','ъ':'ʻ','ь':'','ў':'oʻ','қ':'q','ғ':'gʻ','ҳ':'h','ч':'ch','ш':'sh','а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','з':'z','и':'i','й':'y','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'x','ы':'i','э':'e'}
YOTA = {'е':'ye','ё':'yo','ю':'yu','я':'ya'}

def translit(s):
    if s is None: return None
    s = str(s).replace('\xa0', ' '); out = []; i = 0
    while i < len(s):
        ch = s[i]; low = ch.lower()
        if low == 'ъ' and i+1 < len(s) and s[i+1].lower() in YOTA:
            nxt = s[i+1]; rep = YOTA[nxt.lower()]
            if ch.isupper() or nxt.isupper(): rep = rep[0].upper()+rep[1:]
            out.append(rep); i += 2; continue
        rep = M.get(low)
        if rep is None: out.append(ch); i += 1; continue
        if ch.isupper(): rep = (rep[0].upper()+rep[1:]) if rep else rep
        out.append(rep); i += 1
    return re.sub(r'\s+', ' ', ''.join(out)).strip()

def nc(c): return re.sub(r'\s| ', '', str(c)).strip() if c is not None else None
def fold(s):
    if not s: return ''
    s = ''.join(' ' if ch in APOS else ch for ch in str(s))
    s = unicodedata.normalize('NFKD', s); s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.lower().replace('-', ' ').replace('ts', 's').split())
def tokens(s): return set(fold(s).split())
def sim(a, b):
    ta, tb = tokens(a), tokens(b)
    return len(ta & tb) / max(1, len(ta | tb))

# --- baza ---
base = list(csv.DictReader(open('etl_speciality.csv')))
yrs = {}
for r in csv.DictReader(open('etl_speciality_year.csv')):
    yrs.setdefault(r['speciality_id'], set()).add(int(r['year']))
byce, byname = {}, {}
for r in base:
    byce.setdefault((nc(r['code']), r['education_level']), []).append(r)
    byname.setdefault((fold(r['name_uz']), r['education_level']), []).append(r)
def yl(r):
    y = sorted(yrs.get(r['id'], set()))
    return ','.join(map(str, y)) if y else '—'

# --- statuslar + ranglar ---
GREEN='C6EFCE'; LGREEN='E2EFDA'; YELLOW='FFEB9C'; ORANGE='FCE4D6'; GRAY='D9D9D9'; BLUE='DDEBF7'
STATUS = {
 'YANGI_toza':      ("YO'Q — butunlay yangi",            GREEN),
 'YANGI_kod_qayta': ("YO'Q — yangi (kod qayta ishlatilgan)", LGREEN),
 'BOR_boshqa_yil':  ("BOR — nom bazada (boshqa yil/kod)", YELLOW),
 'NEAR_DUP':        ("BOR — deyarli aynan (near-dup)",    ORANGE),
 'BOR_2026':        ("BOR — shu nom allaqachon 2026'da",  GRAY),
 'CODE_CHANGED':    ("BOR — kod o'zgargan (migration)",   ORANGE),
 'LINK_ONLY':       ("BOR — mavjud, faqat 2026 bog'lash", BLUE),
}
TAVSIYA = {
 'YANGI_toza':      "Qo'shildi (S017)",
 'YANGI_kod_qayta': "Qo'shildi (S017) — kod eski nomdan ozod, yangi yozuv",
 'BOR_boshqa_yil':  "Qo'shildi (S017) — mavjud mutaxassislik, 2026 yozuvi (year-versioned)",
 'NEAR_DUP':        "S017'dan CHIQARILDI — bazada deyarli aynan bor, tekshirish kerak",
 'BOR_2026':        "S017'dan CHIQARILDI — allaqachon 2026'da",
 'CODE_CHANGED':    "S017'dan CHIQARILDI — vazirlik tasdig'i kerak (kod migration)",
 'LINK_ONLY':       "Hech nima kerak emas — allaqachon 2026'da",
}

def classify(note, code, edu, lat):
    n = note.lower()
    codehit = byce.get((code, edu), [])
    namehit = byname.get((fold(lat), edu), [])
    name2026 = any(2026 in yrs.get(x['id'], set()) for x in namehit)
    if 'боғлаш' in n or 'боглаш' in n: return 'LINK_ONLY', codehit, namehit
    if 'ўзгар' in n or 'узгар' in n:   return 'CODE_CHANGED', codehit, namehit
    if namehit and name2026:           return 'BOR_2026', codehit, namehit
    if namehit:                        return 'BOR_boshqa_yil', codehit, namehit
    # nom aniq mos kelmadi — lekin shu KOD band va nomi juda o'xshash bo'lsa near-dup
    if codehit and any(sim(lat, x['name_uz']) >= 0.7 and 2026 in yrs.get(x['id'], set()) for x in codehit):
        return 'NEAR_DUP', codehit, namehit
    if codehit:                        return 'YANGI_kod_qayta', codehit, namehit
    return 'YANGI_toza', codehit, namehit

# --- qatorlarni yig'ish ---
rows = []
for path, edu, dara in [("Бакалавр-кушимча-2026.xlsx","BACHELOR","Bakalavr"),
                        ("Магистр-кушимча-2026.xlsx","MASTER","Magistr")]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.worksheets[0]
    sec = ''
    for r in ws.iter_rows(values_only=True):
        if all(v is None for v in r): continue
        c = list(r) + [None]*6
        code = nc(c[1]); lvl = int(str(c[3])) if str(c[3]).strip().isdigit() else None
        note = str(c[5]).strip() if c[5] else ''
        if lvl == 2: sec = translit(c[2])
        if not note: continue
        lat = translit(c[2])
        k, codehit, namehit = classify(note, code, edu, lat)
        code_ev = ' ; '.join(f"{x['name_uz'][:40]} [{yl(x)}]" for x in codehit) or '— (kod band emas)'
        name_ev = ' ; '.join(f"kod {x['code']} [{yl(x)}]" for x in namehit) or "— (bunday nom yo'q)"
        rows.append({'fayl': f"{dara}-2026", 'daraja': dara, 'code': code,
                     'cyr': str(c[2]).replace('\xa0',' ').strip(), 'lat': lat, 'sec': sec,
                     'note': note, 'kind': k, 'code_ev': code_ev, 'name_ev': name_ev})
    wb.close()

# --- xlsx ---
wb = openpyxl.Workbook()
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF', size=11)
wrap = Alignment(wrap_text=True, vertical='top')

# Varaq 1: Tahlil
ws = wb.active; ws.title = "Tahlil"
head = ['№','Fayl','Daraja','Kod (2026)','Nomi (kirill)','Nomi (lotin)','Bo\'lim (§)',
        'Fayl izohi','HOLAT','Shu KOD bazada nima bor','Shu NOM bazada qayerda','Tavsiya']
ws.append(head)
for cell in ws[1]:
    cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center'); cell.border = border
for i, r in enumerate(rows, 1):
    label, color = STATUS[r['kind']]
    ws.append([i, r['fayl'], r['daraja'], r['code'], r['cyr'], r['lat'], r['sec'],
               r['note'], label, r['code_ev'], r['name_ev'], TAVSIYA[r['kind']]])
    fill = PatternFill('solid', fgColor=color)
    for cell in ws[ws.max_row]:
        cell.border = border; cell.alignment = wrap
    ws.cell(ws.max_row, 9).fill = fill  # HOLAT ustuni rangli
widths = [4,12,9,11,30,30,26,22,30,40,32,42]
for idx, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(idx)].width = w
ws.freeze_panes = 'A2'; ws.auto_filter.ref = f"A1:{get_column_letter(len(head))}{ws.max_row}"

# Varaq 2: Xulosa
from collections import Counter
cat = Counter(r['kind'] for r in rows)
bor = sum(cat[k] for k in ['LINK_ONLY','CODE_CHANGED','BOR_2026','BOR_boshqa_yil','NEAR_DUP'])
yoq = sum(cat[k] for k in ['YANGI_toza','YANGI_kod_qayta'])
xs = wb.create_sheet("Xulosa", 0)
xs['A1'] = "2026 qo'shimcha mutaxassisliklar — bazada BOR / YO'Q tahlili"
xs['A1'].font = Font(bold=True, size=14); xs.merge_cells('A1:C1')
xs.append([]); xs.append(['Jami qator (fayllardagi yangi belgilangan)', 66])
xs.append(["BAZADA YO'Q (rostdan yangi)", yoq]); xs.append(["BAZADA BOR (u yoki bu shaklda)", bor])
xs.append([]); xs.append(['Kategoriya', 'Soni', 'Izoh'])
order = ['YANGI_toza','YANGI_kod_qayta','BOR_boshqa_yil','NEAR_DUP','BOR_2026','CODE_CHANGED','LINK_ONLY']
for k in order:
    xs.append([STATUS[k][0], cat.get(k,0), TAVSIYA[k]])
    xs.cell(xs.max_row,1).fill = PatternFill('solid', fgColor=STATUS[k][1])
for row in xs.iter_rows(min_row=7, max_row=7):
    for cell in row: cell.font = Font(bold=True)
xs.column_dimensions['A'].width = 42; xs.column_dimensions['B'].width = 8; xs.column_dimensions['C'].width = 60
for r in xs.iter_rows():
    for cell in r: cell.alignment = Alignment(wrap_text=True, vertical='center')

out = 'TAHLIL_kushimcha_2026.xlsx'
wb.save(out)
print(f"Yozildi: {out}")
print(f"  BAZADA YO'Q (yangi): {yoq}   BAZADA BOR: {bor}   JAMI: {len(rows)}")
for k in order: print(f"    {STATUS[k][0]:40} {cat.get(k,0)}")
